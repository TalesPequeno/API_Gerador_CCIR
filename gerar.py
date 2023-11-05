import customtkinter as ctk
from tkinter import filedialog
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import openpyxl
import os

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()
        self.todo_sistema()

    def layout_config(self):
        self.title("Gerador de CCIR")
        largura = 700
        altura = 300
        largura_tela = self.winfo_screenwidth()
        altura_tela = self.winfo_screenheight()
        x = (largura_tela - largura) // 2
        y = (altura_tela - altura) // 2
        self.geometry(f"{largura}x{altura}+{x}+{y}")
        self.resizable(width=False, height=False)
        
    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema:", bg_color="transparent", text_color=['#000', "#fff"])
        self.lb_apm.place(x=50, y=230)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["System","Dark", "Light"], command=self.change_apm)
        self.opt_apm.place(x=50, y=260)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=800, height=50, corner_radius=0, bg_color="#3c98e8", fg_color="#3c98e8")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Gerador de CCIR", font=("Century Gothic bold", 24), text_color="#fff")
        title.place(x=250, y=10)

        span = ctk.CTkLabel(self, text="Selecione um arquivo Excel:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        span.place(x=50, y=70)

        linha_frame = ctk.CTkFrame(self)
        linha_frame.place(x=50, y=100)

        self.caminho_entry = ctk.CTkEntry(linha_frame, width=400, font=("Century Gothic bold", 14), state="readonly")
        self.caminho_entry.grid(row=0, column=0)

        botao_buscar = ctk.CTkButton(linha_frame, text="Buscar", font=("Century Gothic bold", 14), command=self.buscar_arquivo)
        botao_buscar.grid(row=0, column=1)

        botao_gerar = ctk.CTkButton(self, text="Gerar", font=("Century Gothic bold", 16), command=self.abrir_chrome)
        botao_gerar.place(x=270, y=200)

        linha_inicio_label = ctk.CTkLabel(self, text="Linha do Início:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        linha_inicio_label.place(x=50, y=160)
        self.linha_inicio_entry = ctk.CTkEntry(self, width=100, font=("Century Gothic bold", 14))
        self.linha_inicio_entry.place(x=160, y=160)

        linha_fim_label = ctk.CTkLabel(self, text="Linha do Término:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        linha_fim_label.place(x=300, y=160)
        self.linha_fim_entry = ctk.CTkEntry(self, width=100, font=("Century Gothic bold", 14))
        self.linha_fim_entry.place(x=430, y=160)

    def buscar_arquivo(self):
        tipos_de_arquivo = [("Arquivos Excel", "*.xls;*.xlsx")]
        arquivo = filedialog.askopenfilename(title="Selecione um arquivo Excel", filetypes=tipos_de_arquivo)
        if arquivo:
            self.caminho_entry.configure(state="normal")
            self.caminho_entry.delete(0, "end")
            self.caminho_entry.insert(0, arquivo)
            self.caminho_entry.configure(state="readonly")

    def abrir_chrome(self):
        url = "https://sncr.serpro.gov.br/ccir/emissao;jsessionid=Nr81me0gvjxVL9rAx4xjgbRj.ccir4?windowId=e0d"
        caminho_arquivo = self.caminho_entry.get()
        linha_inicio = self.linha_inicio_entry.get()
        linha_fim = self.linha_fim_entry.get()

        if not caminho_arquivo:
            messagebox.showerror("Erro", "Por favor, selecione um arquivo Excel antes de clicar em Gerar.")
            return
        
        if not linha_inicio or not linha_fim:
            messagebox.showerror("Erro", "Por favor, preencha os campos de linha de início e término.")
            return
        
        linha_inicio = int(linha_inicio)
        linha_fim = int(linha_fim)

        log_file = open("log.txt", "a")

        wb = openpyxl.load_workbook(caminho_arquivo)
        planilha = wb.active

        for linha in range(linha_inicio, linha_fim + 1):
            dado_coluna_A = planilha.cell(row=linha, column=1).value
            dado_coluna_B = planilha.cell(row=linha, column=2).value
            dado_coluna_C = planilha.cell(row=linha, column=3).value
            dado_coluna_E = planilha.cell(row=linha, column=5).value
            dado_coluna_G = planilha.cell(row=linha, column=7).value

            if dado_coluna_A and dado_coluna_B and dado_coluna_C and dado_coluna_E and dado_coluna_G:
                driver = webdriver.Chrome()
                try:
                    driver.get(url)

                    campo_site = driver.find_element(By.ID, "inputCodigoImovel")
                    campo_site.send_keys(Keys.CONTROL + 'a')
                    campo_site.send_keys(dado_coluna_A)
                    time.sleep(1)

                    campo_site = driver.find_element(By.ID, "selectUfSede")
                    campo_site.click()
                    campo_site.send_keys(dado_coluna_B)
                    campo_site.send_keys(Keys.RETURN)
                    time.sleep(2)

                    campo_site = driver.find_element(By.ID, "selectMunicipioSede")
                    campo_site.click()
                    campo_site.send_keys(dado_coluna_C)
                    campo_site.send_keys(Keys.RETURN)
                    time.sleep(1)

                    campo_site = driver.find_element(By.ID, "inputCpf")
                    campo_site.send_keys(Keys.CONTROL + 'a')
                    campo_site.send_keys(dado_coluna_E)
                    time.sleep(1)

                    # Altera a coordenada do Sou Humano do seu navegador (Google Chrome):
                    coordenada_x = 200
                    coordenada_y = 700

                    action = ActionChains(driver)
                    action.move_by_offset(coordenada_x, coordenada_y).click().perform()
                    time.sleep(3)
                    campo_site.send_keys(Keys.RETURN)

                    campo_site = driver.find_element(By.ID, "formOpcoesCcir:j_idt98")
                    campo_site.click()

                    download_incompleto = True
                    timeout = 10
                    # Altera o caminho de onde é salvo o download
                    caminho_download = "C:/Users/Tales/Downloads"
                    novo_nome = f"CCIR 2023 - {dado_coluna_G}.pdf"

                    while download_incompleto and timeout > 0:
                        time.sleep(1)
                        timeout -= 1

                        arquivos_crdownload = [f for f in os.listdir(caminho_download) if f.lower().endswith(".crdownload")]
                        download_incompleto = bool(arquivos_crdownload)

                    if timeout <= 0:
                        print("O download não foi concluído a tempo.")

                    arquivos_pdf = [f for f in os.listdir(caminho_download) if f.lower().endswith(".pdf")]

                    if arquivos_pdf:
                        arquivo_mais_recente = None
                        data_mais_recente = 0

                        for arquivo in arquivos_pdf:
                            caminho_arquivo = os.path.join(caminho_download, arquivo)
                            data_modificacao = os.path.getmtime(caminho_arquivo)

                            if data_modificacao > data_mais_recente:
                                arquivo_mais_recente = caminho_arquivo
                                data_mais_recente = data_modificacao

                        os.rename(arquivo_mais_recente, os.path.join(caminho_download, novo_nome))
                        log_file.write(f"Arquivo {novo_nome} baixado com sucesso.\n")
                    else:
                        print("Nenhum arquivo PDF encontrado na pasta de download.")
                except Exception as e:
                    print(f"Ocorreu um erro: {str(e)}")
                    log_file.write(f"Ocorreu um erro: {str(e)}\n")
                finally:
                    driver.quit()
            else:
                print(f"Dados na linha {linha} estão incompletos, pulando para a próxima linha.")
                log_file.write(f"Dados na linha {linha} estão incompletos, pulando para a próxima linha.")

        log_file.close()

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

if __name__ == "__main__":
    app = App()
    app.mainloop()
